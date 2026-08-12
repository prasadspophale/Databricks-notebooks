"""Backtest-ready 15m/1h execution scanner with a 2h focus layer.

This intentionally imports the stable VBA-migration implementation and adds:
NSE-aligned 2h candles, ADX/DI, directional focus, confidence decomposition,
nearby-obstacle risk math, and confidence-sorted candidate sheets.
"""

from __future__ import annotations

import argparse
import time
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

import intraday_dashboard_python as base


FOCUS_COLUMNS = [
    "Analysis Time", "Rank", "Stock", "Direction", "Final Recommendation",
    "Confidence", "Priority", "2h Bias", "Timeframe Alignment", "Base Signal", "Entry Status",
    "Base Score", "2h Score", "2h Trend Points", "2h Momentum Points",
    "2h Strength Points", "1h/15m Points", "Risk Points", "2h Close",
    "2h EMA20", "2h EMA50", "2h EMA20 Slope %", "2h RSI", "2h MACD Hist",
    "2h ADX", "2h +DI", "2h -DI", "2h SuperTrend", "Structure",
    "Live Price", "15m Close", "Entry", "Stop Loss", "Target 1", "Target 2",
    "Risk/Reward T1", "Risk/Reward T2", "Nearest Support",
    "Nearest Resistance", "Effective Target", "Effective R/R",
    "Stop Distance %", "Volume", "Avg Volume", "Relative Volume",
    "15m RSI", "1h RSI", "Reason", "Last 15m Candle", "1h Bars", "2h Bars",
    "Data Status",
]

COMPARISON_COLUMNS = [
    "Comparison Signal", "Comparison Score", "Cross-Workbook Agreement",
]

BEST_TRADE_COLUMNS = [
    "Decision", "Stock", "Direction", "Entry", "Stop Loss", "Target 1",
    "Target 2", "Confidence", "Effective R/R", "Stop Distance %",
    "Comparison Signal", "Cross-Workbook Agreement", "Reason",
]


def configure_watchlist(symbols: str | None, symbols_file: Path | None) -> None:
    """Replace the default watchlist from CLI text or a one-column CSV/TXT file."""
    supplied = []
    if symbols:
        supplied.extend(symbols.split(","))
    if symbols_file:
        if not symbols_file.exists():
            raise FileNotFoundError(f"Symbols file not found: {symbols_file}")
        if symbols_file.suffix.lower() == ".csv":
            table = pd.read_csv(symbols_file)
            if table.empty:
                raise ValueError(f"Symbols file is empty: {symbols_file}")
            column = "Symbol" if "Symbol" in table.columns else table.columns[0]
            supplied.extend(table[column].astype(str).tolist())
        else:
            supplied.extend(symbols_file.read_text(encoding="utf-8").replace("\n", ",").split(","))
    if not supplied:
        return
    cleaned = []
    for symbol in supplied:
        value = str(symbol).strip().upper().replace(".NS", "")
        if value and value not in cleaned:
            cleaned.append(value)
    if not cleaned:
        raise ValueError("No valid symbols were supplied")
    base.WATCHLIST = cleaned


def download_extended_history() -> tuple[pd.DataFrame, pd.DataFrame]:
    import yfinance as yf

    frames, failures = [], []
    requests = [("15m", "1mo"), ("1h", "3mo")]
    for number, symbol in enumerate(base.WATCHLIST, 1):
        for interval, period in requests:
            full_symbol = base.yahoo_symbol(symbol)
            try:
                print(f"[{number}/{len(base.WATCHLIST)}] {full_symbol} {interval}")
                raw = yf.download(full_symbol, interval=interval, period=period,
                                  auto_adjust=False, progress=False, threads=False)
                frame = base.normalize_download(raw, symbol, interval)
                if frame.empty:
                    raise ValueError("No valid OHLCV rows returned")
                frames.append(frame)
            except Exception as exc:
                failures.append({"Symbol": full_symbol, "Interval": interval, "Error": str(exc)})
            time.sleep(0.25)
    if not frames:
        raise RuntimeError("Yahoo Finance returned no usable data")
    history = pd.concat(frames, ignore_index=True).sort_values(["Symbol", "Interval", "DateTime"])
    return history.reset_index(drop=True), pd.DataFrame(failures)


def aggregate_2h(hourly: pd.DataFrame, as_of_utc: pd.Timestamp) -> pd.DataFrame:
    """Aggregate complete NSE-aligned 1h rows into complete 2h candles."""
    if hourly.empty:
        return pd.DataFrame(columns=base.HISTORY_COLUMNS + ["SourceBars"])
    work = hourly.copy().sort_values("DateTime")
    utc = pd.to_datetime(work["DateTime"], utc=True)
    ist = utc.dt.tz_convert("Asia/Kolkata")
    session_start = ist.dt.normalize() + pd.Timedelta(hours=9, minutes=15)
    offset = ((ist - session_start).dt.total_seconds() / 60).astype(int)
    valid = (ist.dt.weekday < 5) & (offset >= 0) & (offset <= 300) & (offset % 60 == 0)
    work = work.loc[valid.to_numpy()].copy()
    ist = ist.loc[valid]
    offset = offset.loc[valid]
    work["SessionDate"] = ist.dt.date.to_numpy()
    work["Bucket"] = (offset // 120).to_numpy()
    work = work.loc[work["Bucket"].between(0, 2)]
    rows = []
    for (_, bucket), group in work.groupby(["SessionDate", "Bucket"], sort=True):
        group = group.sort_values("DateTime")
        start_utc = pd.Timestamp(group["DateTime"].iloc[0], tz="UTC")
        if len(group) != 2 or start_utc + pd.Timedelta(hours=2) > as_of_utc:
            continue
        rows.append({
            "Symbol": group["Symbol"].iloc[0], "Interval": "2h",
            "DateTime": group["DateTime"].iloc[0], "Open": group["Open"].iloc[0],
            "High": group["High"].max(), "Low": group["Low"].min(),
            "Close": group["Close"].iloc[-1], "Volume": group["Volume"].sum(),
            "SourceBars": len(group),
        })
    return pd.DataFrame(rows)


def adx_di(data: pd.DataFrame, period: int = 14) -> tuple[float, float, float]:
    high, low, close = data["High"].astype(float), data["Low"].astype(float), data["Close"].astype(float)
    up, down = high.diff(), -low.diff()
    plus_dm = pd.Series(np.where((up > down) & (up > 0), up, 0.0), index=data.index)
    minus_dm = pd.Series(np.where((down > up) & (down > 0), down, 0.0), index=data.index)
    tr = base.true_range(data)
    atrs = tr.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    plus_di = 100 * plus_dm.ewm(alpha=1 / period, adjust=False, min_periods=period).mean() / atrs
    minus_di = 100 * minus_dm.ewm(alpha=1 / period, adjust=False, min_periods=period).mean() / atrs
    dx = 100 * (plus_di - minus_di).abs() / (plus_di + minus_di).replace(0, np.nan)
    adx = dx.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    return float(adx.iloc[-1]), float(plus_di.iloc[-1]), float(minus_di.iloc[-1])


def structure_label(data: pd.DataFrame) -> str:
    recent = data.tail(4)
    if len(recent) < 4:
        return "INSUFFICIENT"
    highs, lows = recent["High"].to_numpy(), recent["Low"].to_numpy()
    if highs[-1] > highs[-2] and lows[-1] > lows[-2]:
        return "HIGHER HIGH / HIGHER LOW"
    if highs[-1] < highs[-2] and lows[-1] < lows[-2]:
        return "LOWER HIGH / LOWER LOW"
    return "MIXED"


def two_hour_metrics(data: pd.DataFrame, direction: str) -> dict:
    if len(data) < 60:
        return {"status": "INSUFFICIENT 2h BARS", "bars": len(data), "score": 0}
    close = data["Close"].astype(float)
    e20, e50 = base.ema(close, 20), base.ema(close, 50)
    rsi = base.rsi(close)
    macd = base.macd_hist(close)
    st = base.supertrend_direction(data)
    adx, plus_di, minus_di = adx_di(data)
    is_long = direction == "LONG"
    regime = (close.iloc[-1] > e20.iloc[-1] > e50.iloc[-1]) if is_long else (close.iloc[-1] < e20.iloc[-1] < e50.iloc[-1])
    slope_ok = (e20.iloc[-1] > e20.iloc[-4]) if is_long else (e20.iloc[-1] < e20.iloc[-4])
    macd_ok = macd > 0 if is_long else macd < 0
    st_ok = st == (1 if is_long else -1)
    rsi_ok = (50 < rsi <= 75) if is_long else (25 <= rsi < 50)
    di_ok = plus_di > minus_di if is_long else minus_di > plus_di
    structure = structure_label(data)
    structure_ok = structure.startswith("HIGHER") if is_long else structure.startswith("LOWER")
    trend_points = (20 if regime else 0) + (5 if slope_ok else 0) + (5 if st_ok else 0)
    momentum_points = (5 if macd_ok else 0) + (5 if rsi_ok else 0)
    strength_points = (5 if adx >= 20 else 0) + (3 if di_ok else 0) + (2 if structure_ok else 0)
    return {
        "status": "OK", "bars": len(data), "score": trend_points + momentum_points + strength_points,
        "bias": "BULLISH" if regime and is_long else "BEARISH" if regime else "NEUTRAL",
        "trend_points": trend_points, "momentum_points": momentum_points,
        "strength_points": strength_points, "close": float(close.iloc[-1]),
        "ema20": float(e20.iloc[-1]), "ema50": float(e50.iloc[-1]),
        "slope_pct": float((e20.iloc[-1] / e20.iloc[-4] - 1) * 100),
        "rsi": rsi, "macd": macd, "adx": adx, "plus_di": plus_di,
        "minus_di": minus_di, "supertrend": "BULLISH" if st == 1 else "BEARISH",
        "structure": structure, "regime": regime,
    }


def obstacle_math(source15: pd.DataFrame, direction: str, entry, stop, target1) -> dict:
    if entry is None or pd.isna(entry) or stop is None or pd.isna(stop):
        return {"support": None, "resistance": None, "effective_target": None,
                "effective_rr": None, "stop_pct": None, "points": 0}
    completed = source15.sort_values("DateTime").tail(40)
    support = float(completed["Low"].iloc[:-1].min())
    resistance = float(completed["High"].iloc[:-1].max())
    risk = abs(float(entry) - float(stop))
    if risk <= 0:
        return {"support": support, "resistance": resistance, "effective_target": None,
                "effective_rr": None, "stop_pct": None, "points": 0}
    if direction == "LONG":
        obstacle = resistance if resistance > float(entry) else float(target1)
        effective_target = min(float(target1), obstacle)
        reward = effective_target - float(entry)
    else:
        obstacle = support if support < float(entry) else float(target1)
        effective_target = max(float(target1), obstacle)
        reward = float(entry) - effective_target
    rr = max(0.0, reward / risk)
    stop_pct = risk / float(entry) * 100
    points = (5 if rr >= 1.5 else 3 if rr >= 1.0 else 0) + (5 if stop_pct <= 2.0 else 3 if stop_pct <= 3.0 else 0)
    return {"support": support, "resistance": resistance, "effective_target": effective_target,
            "effective_rr": rr, "stop_pct": stop_pct, "points": points}


def planned_trade_levels(source15: pd.DataFrame, direction: str,
                         as_of: pd.Timestamp) -> tuple[float, float, float, float] | tuple[None, None, None, None]:
    """Create auditable conditional levels even while the 15m trigger is waiting."""
    d15 = base.completed(source15, 15, as_of)
    if len(d15) < 20:
        return None, None, None, None
    last = d15.iloc[-1]
    a15 = base.atr(d15)
    buffer = max(0.05, float(last["Close"]) * 0.0005)
    if direction == "LONG":
        entry = base.ceil_tick(float(last["High"]) + buffer)
        stop = base.floor_tick(min(float(last["Low"]) - buffer, entry - 1.2 * a15))
        risk = entry - stop
        return entry, stop, base.ceil_tick(entry + 1.5 * risk), base.ceil_tick(entry + 2.5 * risk)
    entry = base.floor_tick(float(last["Low"]) - buffer)
    stop = base.ceil_tick(max(float(last["High"]) + buffer, entry + 1.2 * a15))
    risk = stop - entry
    return entry, stop, base.floor_tick(entry - 1.5 * risk), base.floor_tick(entry - 2.5 * risk)


def build_focus_row(symbol: str, direction: str, history: pd.DataFrame,
                    as_of: pd.Timestamp, two_hour: pd.DataFrame) -> list:
    side = "long" if direction == "LONG" else "short"
    base_row = base.analyze_symbol(symbol, history, side, as_of)
    base_map = dict(zip(base.DASHBOARD_COLUMNS, base_row))
    metrics = two_hour_metrics(two_hour, direction)
    key = base.yahoo_symbol(symbol)
    source15 = history.loc[(history["Symbol"].str.upper().isin([symbol, key])) &
                           (history["Interval"].str.lower() == "15m")]
    entry, stop, target1, target2 = (base_map["Entry"], base_map["Stop Loss"],
                                     base_map["Target 1"], base_map["Target 2"])
    entry_status = base_map["Entry Status"]
    if entry is None or pd.isna(entry):
        entry, stop, target1, target2 = planned_trade_levels(source15, direction, as_of)
        entry_status = "PLANNED - WAIT FOR 15m TRIGGER" if entry is not None else "N/A"
    risk = obstacle_math(source15, direction, entry, stop, target1)
    base_points = round(float(base_map["Score"] or 0) * 0.40, 1)
    confidence = round(float(metrics.get("score", 0)) + base_points + risk["points"], 1)
    aligned = bool(metrics.get("regime")) and base_map["1h Trend"] in ({"BULLISH", "WEAK BULLISH"} if direction == "LONG" else {"BEARISH", "WEAK BEARISH"})
    alignment = "ALIGNED" if aligned else "TIMEFRAME CONFLICT" if metrics.get("regime") else "2h NEUTRAL"
    trade_signals = {"BUY", "STRONG BUY"} if direction == "LONG" else {"SHORT", "STRONG SHORT"}
    if aligned and base_map["Signal"] in trade_signals and risk["effective_rr"] is not None and risk["effective_rr"] >= 1.0:
        recommendation = "CONFIRMED BUY" if direction == "LONG" else "CONFIRMED SELL"
    elif aligned and confidence >= 60:
        recommendation = "WATCH BUY" if direction == "LONG" else "WATCH SELL"
    elif alignment == "TIMEFRAME CONFLICT":
        recommendation = "AVOID - CONFLICT"
    else:
        recommendation = "NO TRADE"
    priority = "HIGH PRIORITY" if confidence >= 80 else "FOCUS" if confidence >= 65 else "WATCH" if confidence >= 50 else "IGNORE"
    avg_volume = base_map["Avg Volume"] or 0
    rel_volume = float(base_map["Volume"] or 0) / float(avg_volume) if avg_volume else None
    reason = f"{metrics.get('bias', 'NO 2h BIAS')}; {base_map['Reason']}"
    return [
        datetime.now(), None, symbol, direction, recommendation, confidence, priority,
        metrics.get("bias", "N/A"), alignment, base_map["Signal"], entry_status, base_map["Score"],
        metrics.get("score", 0), metrics.get("trend_points", 0), metrics.get("momentum_points", 0),
        metrics.get("strength_points", 0), base_points, risk["points"], metrics.get("close"),
        metrics.get("ema20"), metrics.get("ema50"), metrics.get("slope_pct"), metrics.get("rsi"),
        metrics.get("macd"), metrics.get("adx"), metrics.get("plus_di"), metrics.get("minus_di"),
        metrics.get("supertrend"), metrics.get("structure"), base_map["Live Price"],
        base_map["15m Close"], entry, stop, target1, target2, 1.5, 2.5,
        risk["support"], risk["resistance"], risk["effective_target"], risk["effective_rr"],
        risk["stop_pct"], base_map["Volume"], base_map["Avg Volume"], rel_volume,
        base_map["15m RSI"], base_map["1h RSI"], reason, base_map["15m Candle Time"],
        base_map["1h Bars"], metrics.get("bars", 0), base_map["Data Status"],
    ]


def build_focus(history: pd.DataFrame, replay: bool) -> tuple[pd.DataFrame, pd.DataFrame]:
    maximum = pd.to_datetime(history["DateTime"], utc=True).max()
    as_of = maximum + pd.Timedelta(minutes=1) if replay else pd.Timestamp.now(tz="UTC")
    focus_rows, history2 = [], []
    for symbol in base.WATCHLIST:
        key = base.yahoo_symbol(symbol)
        hourly = history.loc[(history["Symbol"].str.upper().isin([symbol, key])) &
                             (history["Interval"].str.lower().isin(["1h", "60m"]))]
        bars2 = aggregate_2h(hourly, as_of)
        if not bars2.empty:
            history2.append(bars2)
        focus_rows.append(build_focus_row(symbol, "LONG", history, as_of, bars2))
        focus_rows.append(build_focus_row(symbol, "SHORT", history, as_of, bars2))
    focus = pd.DataFrame(focus_rows, columns=FOCUS_COLUMNS)
    focus = focus.sort_values(["Confidence", "Base Score"], ascending=False).reset_index(drop=True)
    focus["Rank"] = np.arange(1, len(focus) + 1)
    return focus, pd.concat(history2, ignore_index=True) if history2 else pd.DataFrame()


def load_comparison_signals(workbook: Path | None) -> pd.DataFrame:
    """Read the earlier 15m/1h dashboards and return one row per symbol/side."""
    if workbook is None:
        return pd.DataFrame(columns=["Stock", "Direction", *COMPARISON_COLUMNS])
    if not workbook.exists():
        raise FileNotFoundError(f"Comparison workbook not found: {workbook}")
    frames = []
    for sheet_name, direction, score_name in (
        ("Long Dashboard", "LONG", "Bull Score"),
        ("Short Dashboard", "SHORT", "Bear Score"),
    ):
        try:
            dashboard = pd.read_excel(workbook, sheet_name=sheet_name, engine="openpyxl")
        except ValueError:
            continue
        required = {"Stock", "Signal", score_name}
        if not required.issubset(dashboard.columns):
            missing = sorted(required.difference(dashboard.columns))
            raise ValueError(f"{sheet_name} is missing columns: {missing}")
        view = dashboard[["Stock", "Signal", score_name]].copy()
        view["Stock"] = view["Stock"].astype(str).str.upper().str.replace(".NS", "", regex=False)
        view["Direction"] = direction
        view = view.rename(columns={
            "Signal": "Comparison Signal", score_name: "Comparison Score",
        })
        expected = {"BUY", "STRONG BUY"} if direction == "LONG" else {"SHORT", "STRONG SHORT"}
        view["Cross-Workbook Agreement"] = view["Comparison Signal"].isin(expected)
        frames.append(view)
    if not frames:
        raise ValueError("Comparison workbook has neither Long Dashboard nor Short Dashboard")
    return pd.concat(frames, ignore_index=True)


def apply_workbook_comparison(focus: pd.DataFrame, comparison: pd.DataFrame) -> pd.DataFrame:
    """Attach like-for-like symbol/direction agreement without changing model confidence."""
    output = focus.copy()
    output["Stock"] = output["Stock"].astype(str).str.upper().str.replace(".NS", "", regex=False)
    if comparison.empty:
        output["Comparison Signal"] = "NOT PROVIDED"
        output["Comparison Score"] = np.nan
        output["Cross-Workbook Agreement"] = False
        return output
    return output.merge(comparison, on=["Stock", "Direction"], how="left").assign(
        **{
            "Comparison Signal": lambda data: data["Comparison Signal"].fillna("MISSING"),
            "Cross-Workbook Agreement": lambda data: data["Cross-Workbook Agreement"].fillna(False),
        }
    )


def select_single_trade(focus: pd.DataFrame, require_comparison: bool) -> pd.DataFrame:
    """Return exactly one executable trade, or an explicit NO TRADE decision."""
    confirmed = focus["Final Recommendation"].isin(["CONFIRMED BUY", "CONFIRMED SELL"])
    valid_levels = focus[["Entry", "Stop Loss", "Target 1", "Target 2"]].notna().all(axis=1)
    risk_ok = focus["Stop Distance %"].le(2.0) & focus["Effective R/R"].ge(1.0)
    data_ok = focus["Data Status"].eq("OK") & focus["Timeframe Alignment"].eq("ALIGNED")
    agreement_ok = focus["Cross-Workbook Agreement"] if require_comparison else True
    eligible = focus.loc[confirmed & valid_levels & risk_ok & data_ok & agreement_ok].copy()
    if eligible.empty:
        reason = "No candidate passed confirmation, alignment, data, <=2% stop, and >=1.0 effective R/R"
        if require_comparison:
            reason += " with agreement in both workbooks"
        return pd.DataFrame([{
            "Decision": "NO TRADE", "Stock": None, "Direction": None,
            "Entry": None, "Stop Loss": None, "Target 1": None, "Target 2": None,
            "Confidence": None, "Effective R/R": None, "Stop Distance %": None,
            "Comparison Signal": None, "Cross-Workbook Agreement": False,
            "Reason": reason,
        }], columns=BEST_TRADE_COLUMNS)
    eligible["Expected Reward T2 %"] = np.where(
        eligible["Direction"].eq("LONG"),
        (eligible["Target 2"] / eligible["Entry"] - 1) * 100,
        (eligible["Entry"] / eligible["Target 2"] - 1) * 100,
    )
    best = eligible.sort_values(
        ["Confidence", "Effective R/R", "Expected Reward T2 %", "Stop Distance %"],
        ascending=[False, False, False, True],
    ).iloc[0]
    decision = "BUY" if best["Direction"] == "LONG" else "SELL"
    return pd.DataFrame([{
        "Decision": decision, "Stock": best["Stock"], "Direction": best["Direction"],
        "Entry": best["Entry"], "Stop Loss": best["Stop Loss"],
        "Target 1": best["Target 1"], "Target 2": best["Target 2"],
        "Confidence": best["Confidence"], "Effective R/R": best["Effective R/R"],
        "Stop Distance %": best["Stop Distance %"],
        "Comparison Signal": best["Comparison Signal"],
        "Cross-Workbook Agreement": bool(best["Cross-Workbook Agreement"]),
        "Reason": "Highest-ranked fully confirmed, risk-filtered cross-workbook candidate",
    }], columns=BEST_TRADE_COLUMNS)


def save_focus_workbook(history: pd.DataFrame, history2: pd.DataFrame, focus: pd.DataFrame,
                        best_trade: pd.DataFrame, failures: pd.DataFrame, output: Path) -> None:
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    long = focus.loc[focus["Direction"] == "LONG"].copy()
    short = focus.loc[focus["Direction"] == "SHORT"].copy()
    action_rank = focus["Final Recommendation"].map({
        "CONFIRMED BUY": 4, "CONFIRMED SELL": 4, "WATCH BUY": 3,
        "WATCH SELL": 3, "NO TRADE": 1, "AVOID - CONFLICT": 0,
    }).fillna(0)
    best = (focus.assign(_ActionRank=action_rank)
            .sort_values(["_ActionRank", "Confidence"], ascending=False)
            .groupby("Stock", as_index=False).first()
            .drop(columns="_ActionRank").sort_values("Confidence", ascending=False))
    best["Rank"] = np.arange(1, len(best) + 1)
    best = best[FOCUS_COLUMNS + COMPARISON_COLUMNS]
    confidence_ok = np.isclose(focus["Confidence"], focus["2h Score"] + focus["1h/15m Points"] + focus["Risk Points"]).all()
    level_rows = focus.dropna(subset=["Entry", "Stop Loss", "Target 1", "Target 2"])
    long_levels = level_rows.loc[level_rows["Direction"] == "LONG"]
    short_levels = level_rows.loc[level_rows["Direction"] == "SHORT"]
    long_ok = ((long_levels["Stop Loss"] < long_levels["Entry"]) &
               (long_levels["Entry"] < long_levels["Target 1"]) &
               (long_levels["Target 1"] < long_levels["Target 2"])).all()
    short_ok = ((short_levels["Stop Loss"] > short_levels["Entry"]) &
                (short_levels["Entry"] > short_levels["Target 1"]) &
                (short_levels["Target 1"] > short_levels["Target 2"])).all()
    rr_ok = (level_rows["Effective R/R"].dropna() >= 0).all()
    checks = pd.DataFrame([
        ["Symbols expected", len(base.WATCHLIST), len(best), "OK" if len(best) == len(base.WATCHLIST) else "FAIL"],
        ["2h bars available", len(base.WATCHLIST), int((best["2h Bars"] >= 60).sum()), "OK" if (best["2h Bars"] >= 60).all() else "REVIEW"],
        ["Valid data status", len(base.WATCHLIST), int((best["Data Status"] == "OK").sum()), "OK" if (best["Data Status"] == "OK").all() else "REVIEW"],
        ["Confidence decomposition", 1, int(confidence_ok), "OK" if confidence_ok else "FAIL"],
        ["Long entry/stop/target direction", 1, int(long_ok), "OK" if long_ok else "FAIL"],
        ["Short entry/stop/target direction", 1, int(short_ok), "OK" if short_ok else "FAIL"],
        ["Effective R/R non-negative", 1, int(rr_ok), "OK" if rr_ok else "FAIL"],
        ["Confirmed candidates", None, int(best["Final Recommendation"].str.startswith("CONFIRMED").sum()), "INFO"],
    ], columns=["Check", "Expected", "Actual", "Status"])
    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        history.to_excel(writer, sheet_name="History1", index=False)
        history2.to_excel(writer, sheet_name="History2H", index=False)
        best.to_excel(writer, sheet_name="Today Focus", index=False)
        best_trade.to_excel(writer, sheet_name="Best Trade", index=False)
        long.to_excel(writer, sheet_name="Long Candidates", index=False)
        short.to_excel(writer, sheet_name="Short Candidates", index=False)
        checks.to_excel(writer, sheet_name="Checks", index=False)
        if not failures.empty:
            failures.to_excel(writer, sheet_name="Download Errors", index=False)
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.font = Font(color="FFFFFF", bold=True)
            for cells in ws.columns:
                letter = get_column_letter(cells[0].column)
                ws.column_dimensions[letter].width = min(max(len(str(c.value or "")) for c in cells[:100]) + 2, 42)
            if ws.title in {"Today Focus", "Long Candidates", "Short Candidates"}:
                conf_col = FOCUS_COLUMNS.index("Confidence") + 1
                conf_letter = get_column_letter(conf_col)
                ws.conditional_formatting.add(f"{conf_letter}2:{conf_letter}{ws.max_row}",
                    CellIsRule(operator="greaterThanOrEqual", formula=["80"], fill=PatternFill("solid", fgColor="C6EFCE")))
                ws.column_dimensions["E"].width = 22
                ws.column_dimensions[get_column_letter(FOCUS_COLUMNS.index("Reason") + 1)].width = 48
    print(f"Workbook saved: {output}")


def main() -> None:
    parser = argparse.ArgumentParser(description="15m/1h/2h confidence-ranked intraday focus scanner")
    parser.add_argument("--history-file", type=Path, help="Replay an existing History1 workbook")
    parser.add_argument("--comparison-file", type=Path,
                        help="Earlier workbook containing Long Dashboard and Short Dashboard")
    parser.add_argument("--symbols", help="Comma-separated symbols; overrides the default watchlist")
    parser.add_argument("--symbols-file", type=Path,
                        help="TXT or CSV watchlist; CSV uses Symbol column or its first column")
    parser.add_argument("--output", type=Path, default=Path("intraday_focus_15m_1h_2h.xlsx"))
    args = parser.parse_args()
    configure_watchlist(args.symbols, args.symbols_file)
    history, failures = (base.load_history(args.history_file), pd.DataFrame()) if args.history_file else download_extended_history()
    focus, history2 = build_focus(history, replay=bool(args.history_file))
    comparison = load_comparison_signals(args.comparison_file)
    focus = apply_workbook_comparison(focus, comparison)
    best_trade = select_single_trade(focus, require_comparison=args.comparison_file is not None)
    save_focus_workbook(history, history2, focus, best_trade, failures, args.output)
    print("\nSingle-trade decision:")
    print(best_trade.to_string(index=False))
    print(focus[["Rank", "Stock", "Direction", "Final Recommendation", "Confidence", "Priority"]].head(15).to_string(index=False))


if __name__ == "__main__":
    main()
