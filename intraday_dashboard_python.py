"""Python migration of prasadspophale/ExcelVBACode intraday VBA scanners.

Downloads Yahoo Finance 15m/1h history, calculates the VBA strategy indicators,
and creates History1, Short Dashboard, and Long Dashboard worksheets.
"""

from __future__ import annotations

import argparse
import math
import time
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd


WATCHLIST = [
    "IREDA", "BDL", "TATAELXSI", "KAYNES", "NTPCGREEN", "NETWEB",
    "IRFC", "NTPC", "BEL", "JIOFIN", "HUDCO", "RELIANCE", "TATAPOWER",
    "TCS", "TEJASNET", "JSWENERGY", "NHPC", "TATATECH", "ADANIPORTS",
    "COCHINSHIP", "HAL", "AFFLE", "INOXGREEN", "KELLTONTEC", "POWERGRID",
    "ZODIAC", "ORBTEXP", "HDFCSML250", "VISHNU", "ICICIBANK",
]

HISTORY_COLUMNS = ["Symbol", "Interval", "DateTime", "Open", "High", "Low", "Close", "Volume"]
DASHBOARD_COLUMNS = [
    "Analysis Time", "Stock", "Signal", "Entry Status", "Score", "Reason",
    "15m Candle Time", "Live Price", "15m Close", "Entry", "Stop Loss",
    "Target 1", "Target 2", "Risk/Reward T1", "Risk/Reward T2", "1h Trend",
    "1h RSI", "1h EMA20", "1h EMA50", "1h MACD Hist", "15m RSI",
    "15m EMA9", "15m EMA20", "15m VWAP", "15m ATR", "15m MACD Hist",
    "Volume", "Avg Volume", "15m Bars", "1h Bars", "Data Status",
]


def yahoo_symbol(symbol: str) -> str:
    symbol = symbol.strip().upper()
    return symbol if symbol.endswith(".NS") or symbol.startswith("^") else f"{symbol}.NS"


def normalize_download(raw: pd.DataFrame, symbol: str, interval: str) -> pd.DataFrame:
    if raw.empty:
        return pd.DataFrame(columns=HISTORY_COLUMNS)
    if isinstance(raw.columns, pd.MultiIndex):
        raw.columns = raw.columns.get_level_values(0)
    raw = raw.reset_index()
    date_col = "Datetime" if "Datetime" in raw.columns else "Date"
    dates = pd.to_datetime(raw[date_col], utc=True).dt.tz_localize(None)
    out = pd.DataFrame({
        "Symbol": yahoo_symbol(symbol), "Interval": interval, "DateTime": dates,
        "Open": pd.to_numeric(raw["Open"], errors="coerce"),
        "High": pd.to_numeric(raw["High"], errors="coerce"),
        "Low": pd.to_numeric(raw["Low"], errors="coerce"),
        "Close": pd.to_numeric(raw["Close"], errors="coerce"),
        "Volume": pd.to_numeric(raw["Volume"], errors="coerce").fillna(0),
    })
    return out.dropna(subset=["DateTime", "Open", "High", "Low", "Close"])


def download_history() -> tuple[pd.DataFrame, pd.DataFrame]:
    import yfinance as yf

    frames, failures = [], []
    requests = [("15m", "5d"), ("1h", "3mo")]
    for number, symbol in enumerate(WATCHLIST, 1):
        for interval, period in requests:
            full_symbol = yahoo_symbol(symbol)
            try:
                print(f"[{number}/{len(WATCHLIST)}] {full_symbol} {interval}")
                raw = yf.download(full_symbol, interval=interval, period=period,
                                  auto_adjust=False, progress=False, threads=False)
                frame = normalize_download(raw, symbol, interval)
                if frame.empty:
                    raise ValueError("No valid OHLCV rows returned")
                frames.append(frame)
            except Exception as exc:
                failures.append({"Symbol": full_symbol, "Interval": interval, "Error": str(exc)})
            time.sleep(0.25)
    if not frames:
        raise RuntimeError("Yahoo Finance returned no usable data")
    history = pd.concat(frames, ignore_index=True).sort_values(["Symbol", "Interval", "DateTime"])
    return history.reset_index(drop=True), pd.DataFrame(failures)


def load_history(path: Path) -> pd.DataFrame:
    frame = pd.read_excel(path, sheet_name="History1")
    missing = set(HISTORY_COLUMNS) - set(frame.columns)
    if missing:
        raise ValueError(f"History1 is missing columns: {', '.join(sorted(missing))}")
    frame = frame[HISTORY_COLUMNS].copy()
    frame["DateTime"] = pd.to_datetime(frame["DateTime"], errors="coerce")
    return frame.dropna(subset=["DateTime", "Open", "High", "Low", "Close"])


def ema(values: pd.Series, period: int) -> pd.Series:
    return values.astype(float).ewm(span=period, adjust=False, min_periods=period).mean()


def rsi(values: pd.Series, period: int = 14) -> float:
    delta = values.astype(float).diff()
    gain = delta.clip(lower=0).ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    loss = (-delta.clip(upper=0)).ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    rs = gain / loss.replace(0, np.nan)
    result = 100 - 100 / (1 + rs)
    if loss.iloc[-1] == 0 and gain.iloc[-1] > 0:
        return 100.0
    return float(result.iloc[-1])


def true_range(data: pd.DataFrame) -> pd.Series:
    previous = data["Close"].shift(1)
    return pd.concat([(data["High"] - data["Low"]).abs(),
                      (data["High"] - previous).abs(),
                      (data["Low"] - previous).abs()], axis=1).max(axis=1)


def atr(data: pd.DataFrame, period: int = 14) -> float:
    return float(true_range(data).ewm(alpha=1 / period, adjust=False, min_periods=period).mean().iloc[-1])


def macd_hist(values: pd.Series) -> float:
    fast, slow = ema(values, 12), ema(values, 26)
    line = fast - slow
    signal = line.ewm(span=9, adjust=False, min_periods=9).mean()
    return float((line - signal).iloc[-1])


def supertrend_direction(data: pd.DataFrame, period: int = 10, multiplier: float = 3.0) -> int:
    tr = true_range(data)
    atrs = tr.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    middle = (data["High"] + data["Low"]) / 2
    upper, lower = middle + multiplier * atrs, middle - multiplier * atrs
    final_upper, final_lower = upper.copy(), lower.copy()
    direction = pd.Series(1, index=data.index, dtype=int)
    for i in range(1, len(data)):
        final_upper.iloc[i] = upper.iloc[i] if (upper.iloc[i] < final_upper.iloc[i - 1] or
                                                data["Close"].iloc[i - 1] > final_upper.iloc[i - 1]) else final_upper.iloc[i - 1]
        final_lower.iloc[i] = lower.iloc[i] if (lower.iloc[i] > final_lower.iloc[i - 1] or
                                                data["Close"].iloc[i - 1] < final_lower.iloc[i - 1]) else final_lower.iloc[i - 1]
        if data["Close"].iloc[i] > final_upper.iloc[i - 1]:
            direction.iloc[i] = 1
        elif data["Close"].iloc[i] < final_lower.iloc[i - 1]:
            direction.iloc[i] = -1
        else:
            direction.iloc[i] = direction.iloc[i - 1]
    return int(direction.iloc[-1])


def session_vwap(data: pd.DataFrame) -> float:
    local = data["DateTime"].dt.tz_localize("UTC").dt.tz_convert("Asia/Kolkata")
    latest_date = local.iloc[-1].date()
    session = data.loc[local.dt.date == latest_date]
    typical = (session["High"] + session["Low"] + session["Close"]) / 3
    volume = session["Volume"].clip(lower=0)
    return float((typical * volume).sum() / volume.sum()) if volume.sum() else float(session["Close"].iloc[-1])


def completed(data: pd.DataFrame, interval_minutes: int, as_of_utc: pd.Timestamp) -> pd.DataFrame:
    data = data.sort_values("DateTime").copy()
    utc = pd.to_datetime(data["DateTime"], utc=True)
    ist = utc.dt.tz_convert("Asia/Kolkata")
    start = ist.dt.normalize() + pd.Timedelta(hours=9, minutes=15)
    offset = (ist - start).dt.total_seconds() / 60
    maximum = 360 if interval_minutes == 15 else 300
    aligned = ((offset >= 0) & (offset <= maximum) & (offset % interval_minutes == 0) &
               (ist.dt.weekday < 5))
    closed = utc + pd.Timedelta(minutes=interval_minutes) <= as_of_utc
    return data.loc[aligned.to_numpy() & closed.to_numpy()].reset_index(drop=True)


def floor_tick(value: float) -> float:
    return math.floor(value / 0.05 + 1e-7) * 0.05


def ceil_tick(value: float) -> float:
    return math.ceil(value / 0.05 - 1e-7) * 0.05


def freshness(last_utc: pd.Timestamp, as_of_utc: pd.Timestamp) -> str:
    age = (as_of_utc - (last_utc + pd.Timedelta(minutes=15))).total_seconds() / 60
    now_ist = as_of_utc.tz_convert("Asia/Kolkata")
    if age < 0:
        return "INCOMPLETE CANDLE"
    if now_ist.weekday() < 5 and (9, 15) <= (now_ist.hour, now_ist.minute) <= (15, 45) and age > 45:
        return "STALE 15m DATA"
    return "OK"


def error_row(symbol: str, status: str, bars15: int, bars60: int) -> list:
    row = [None] * 31
    row[0:6] = [datetime.now(), symbol, "NO TRADE", "N/A", 0, status]
    row[28:31] = [bars15, bars60, status]
    return row


def analyze_symbol(symbol: str, history: pd.DataFrame, side: str, as_of_utc: pd.Timestamp) -> list:
    key = yahoo_symbol(symbol)
    source = history.loc[history["Symbol"].str.upper().isin([symbol, key])]
    raw15, raw60 = source.loc[source["Interval"].str.lower() == "15m"], source.loc[source["Interval"].str.lower().isin(["1h", "60m"])]
    if raw15.empty or raw60.empty:
        return error_row(symbol, "15m DATA MISSING" if raw15.empty else "1h DATA MISSING", len(raw15), len(raw60))
    live_price = float(raw15.sort_values("DateTime")["Close"].iloc[-1])
    d15, d60 = completed(raw15, 15, as_of_utc), completed(raw60, 60, as_of_utc)
    if len(d15) < 60 or len(d60) < 60:
        return error_row(symbol, "INSUFFICIENT COMPLETED CANDLES", len(d15), len(d60))

    e9, e20 = ema(d15["Close"], 9), ema(d15["Close"], 20)
    e20h, e50h = ema(d60["Close"], 20), ema(d60["Close"], 50)
    r15, r60 = rsi(d15["Close"]), rsi(d60["Close"])
    a15, m15, m60 = atr(d15), macd_hist(d15["Close"]), macd_hist(d60["Close"])
    vw, st60 = session_vwap(d15), supertrend_direction(d60)
    last, prev = d15.iloc[-1], d15.iloc[-2]
    positive_prior = d15["Volume"].iloc[-21:-1]
    positive_prior = positive_prior[positive_prior > 0]
    avg_volume = float(positive_prior.mean()) if len(positive_prior) else 0.0
    is_long = side == "long"

    regime = (d60["Close"].iloc[-1] > e20h.iloc[-1] > e50h.iloc[-1]) if is_long else (d60["Close"].iloc[-1] < e20h.iloc[-1] < e50h.iloc[-1])
    vwap_ok = last["Close"] > vw if is_long else last["Close"] < vw
    ema_ok = e9.iloc[-1] > e20.iloc[-1] if is_long else e9.iloc[-1] < e20.iloc[-1]
    primary = last["Close"] > prev["High"] if is_long else last["Close"] < prev["Low"]
    rejection = ((last["Low"] <= e20.iloc[-1] and last["Close"] > e20.iloc[-1] and last["Close"] > last["Open"])
                 if is_long else
                 (last["High"] >= e20.iloc[-1] and last["Close"] < e20.iloc[-1] and last["Close"] < last["Open"]))
    trigger = vwap_ok and ema_ok and (primary or rejection)

    score = 0
    score += 20 if regime else 0
    score += 10 if ((e20h.iloc[-1] > e20h.iloc[-4]) == is_long) else 0
    score += 10 if ((m60 > 0) == is_long) else 0
    score += 10 if st60 == (1 if is_long else -1) else 0
    score += 10 if ((50 < r60 <= 75) if is_long else (25 <= r60 < 50)) else 0
    score += 10 if vwap_ok else 0
    score += 10 if ema_ok else 0
    score += 5 if ((m15 > 0) == is_long) else 0
    score += 5 if ((55 < r15 <= 75) if is_long else (25 <= r15 < 45)) else 0
    score += 5 if primary or rejection else 0
    score += 5 if avg_volume > 0 and last["Volume"] > 1.2 * avg_volume else 0

    status = freshness(pd.Timestamp(d15["DateTime"].iloc[-1], tz="UTC"), as_of_utc)
    signal, reason = "NO TRADE", status
    if status == "OK":
        if not regime:
            reason = f"1h {'bullish' if is_long else 'bearish'} regime not confirmed"
        elif (is_long and r15 > 75) or (not is_long and r15 < 25):
            reason = f"15m RSI is {'overbought' if is_long else 'oversold'}; avoid chasing"
        elif not trigger:
            signal, reason = "WATCH", f"1h {'bullish' if is_long else 'bearish'}; waiting for 15m {'breakout/pullback' if is_long else 'breakdown/rejection'}"
        elif score >= 75:
            signal, reason = ("STRONG BUY" if is_long else "STRONG SHORT"), f"1h {'bullish' if is_long else 'bearish'} and 15m {'buy' if is_long else 'short'} trigger confirmed"
        elif score >= 60:
            signal, reason = ("BUY" if is_long else "SHORT"), f"{'Buy' if is_long else 'Short'} conditions confirmed"
        else:
            signal, reason = "WATCH", f"{'Bullish' if is_long else 'Bearish'} setup has insufficient confirmation"

    entry = stop = t1 = t2 = rr1 = rr2 = None
    entry_status = "N/A"
    if signal in {"BUY", "STRONG BUY", "SHORT", "STRONG SHORT"}:
        buffer = max(0.05, float(last["Close"]) * 0.0005)
        if is_long:
            entry = ceil_tick(float(last["High"]) + buffer)
            stop = floor_tick(min(float(last["Low"]) - buffer, entry - 1.2 * a15))
            risk = entry - stop
            t1, t2 = ceil_tick(entry + 1.5 * risk), ceil_tick(entry + 2.5 * risk)
            entry_status = "WAITING" if live_price < entry else ("ENTRY TRIGGERED" if live_price <= entry + 0.25 * a15 else "ENTRY MISSED")
        else:
            entry = floor_tick(float(last["Low"]) - buffer)
            stop = ceil_tick(max(float(last["High"]) + buffer, entry + 1.2 * a15))
            risk = stop - entry
            t1, t2 = floor_tick(entry - 1.5 * risk), floor_tick(entry - 2.5 * risk)
            entry_status = "WAITING" if live_price > entry else ("ENTRY TRIGGERED" if live_price >= entry - 0.25 * a15 else "ENTRY MISSED")
        rr1, rr2 = 1.5, 2.5

    if regime and st60 == (1 if is_long else -1):
        trend = "BULLISH" if is_long else "BEARISH"
    elif (d60["Close"].iloc[-1] > e20h.iloc[-1]) == is_long:
        trend = "WEAK BULLISH" if is_long else "WEAK BEARISH"
    else:
        trend = "NOT BULLISH" if is_long else "NOT BEARISH"
    candle_ist = pd.Timestamp(last["DateTime"], tz="UTC").tz_convert("Asia/Kolkata").tz_localize(None)
    return [datetime.now(), symbol, signal, entry_status, score, reason, candle_ist,
            round(live_price, 2), round(float(last["Close"]), 2), entry, stop, t1, t2,
            rr1, rr2, trend, round(r60, 2), round(float(e20h.iloc[-1]), 2),
            round(float(e50h.iloc[-1]), 2), round(m60, 3), round(r15, 2),
            round(float(e9.iloc[-1]), 2), round(float(e20.iloc[-1]), 2), round(vw, 2),
            round(a15, 2), round(m15, 3), int(last["Volume"]), int(avg_volume),
            len(d15), len(d60), status]


def create_dashboards(history: pd.DataFrame, historical_replay: bool = False) -> tuple[pd.DataFrame, pd.DataFrame]:
    maximum = pd.to_datetime(history["DateTime"], utc=True).max()
    # In an archived Yahoo snapshot, the newest 15m row is the live/incomplete
    # candle, just as in VBA. Replay one minute into that candle so it supplies
    # Live Price but is excluded from completed-candle indicators and triggers.
    as_of = (maximum + pd.Timedelta(minutes=1) if historical_replay
             else pd.Timestamp.now(tz="UTC"))
    short_rows = [analyze_symbol(s, history, "short", as_of) for s in WATCHLIST]
    long_rows = [analyze_symbol(s, history, "long", as_of) for s in WATCHLIST]
    short = pd.DataFrame(short_rows, columns=DASHBOARD_COLUMNS).rename(columns={"Score": "Bear Score"})
    long = pd.DataFrame(long_rows, columns=DASHBOARD_COLUMNS).rename(columns={"Score": "Bull Score"})
    return short.sort_values("Bear Score", ascending=False), long.sort_values("Bull Score", ascending=False)


def save_workbook(history: pd.DataFrame, short: pd.DataFrame, long: pd.DataFrame,
                  failures: pd.DataFrame, output: Path) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        history.to_excel(writer, sheet_name="History1", index=False)
        short.to_excel(writer, sheet_name="Short Dashboard", index=False)
        long.to_excel(writer, sheet_name="Long Dashboard", index=False)
        if not failures.empty:
            failures.to_excel(writer, sheet_name="Download Errors", index=False)
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.font = Font(color="FFFFFF", bold=True)
            for column_cells in ws.columns:
                letter = get_column_letter(column_cells[0].column)
                width = min(max(len(str(c.value or "")) for c in column_cells[:200]) + 2, 48)
                ws.column_dimensions[letter].width = width
            if ws.title == "History1":
                for cell in ws["C"][1:]: cell.number_format = "dd-mmm-yyyy hh:mm"
                for col in "DEFG":
                    for cell in ws[col][1:]: cell.number_format = "0.00"
                for cell in ws["H"][1:]: cell.number_format = "#,##0"
            elif "Dashboard" in ws.title:
                for col in ("A", "G"):
                    for cell in ws[col][1:]: cell.number_format = "dd-mmm-yyyy hh:mm"
                for col in range(8, 27):
                    for cell in ws.cell(row=2, column=col).offset(row=0, column=0).parent.iter_cols(min_col=col, max_col=col, min_row=2, max_row=ws.max_row):
                        for item in cell: item.number_format = "0.00"
                ws.column_dimensions["F"].width = 48
    print(f"Workbook saved: {output}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Migrate the intraday VBA dashboard to Python")
    parser.add_argument("--history-file", type=Path, help="Use an existing workbook History1 sheet instead of Yahoo")
    parser.add_argument("--output", type=Path, default=Path("intraday_python_output.xlsx"))
    args = parser.parse_args()
    if args.history_file:
        history, failures = load_history(args.history_file), pd.DataFrame()
    else:
        history, failures = download_history()
    short, long = create_dashboards(history, historical_replay=bool(args.history_file))
    save_workbook(history, short, long, failures, args.output)
    print(short[["Stock", "Signal", "Bear Score"]].head(10).to_string(index=False))
    print(long[["Stock", "Signal", "Bull Score"]].head(10).to_string(index=False))


if __name__ == "__main__":
    main()

